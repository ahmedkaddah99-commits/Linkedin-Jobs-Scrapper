from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Iterable, Mapping
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse


LOGGER = logging.getLogger(__name__)


def compact_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def normalize_title_company_part(value: Any) -> str:
    return compact_whitespace(str(value or "")).lower()


def title_company_signature(title: Any, company: Any) -> str:
    title_part = normalize_title_company_part(title)
    company_part = normalize_title_company_part(company)
    if not title_part or not company_part:
        return ""
    return f"{title_part}||{company_part}"


def canonicalize_url(raw_url: str) -> str:
    value = compact_whitespace(raw_url)
    if not value:
        return ""

    try:
        parsed = urlparse(value)
    except Exception:
        return ""

    scheme = (parsed.scheme or "").lower()
    if scheme not in {"http", "https"}:
        return ""

    netloc = (parsed.netloc or "").lower()
    if netloc.startswith("www."):
        netloc = netloc[4:]
    if not netloc:
        return ""

    path = re.sub(r"/{2,}", "/", parsed.path or "").rstrip("/")
    if not path:
        path = "/"

    cleaned_query = []
    for key, value in parse_qsl(parsed.query, keep_blank_values=True):
        key_lower = key.lower()
        if key_lower.startswith("utm_") or key_lower in {
            "trk",
            "trackingid",
            "refid",
            "ref",
            "src",
            "source",
        }:
            continue
        cleaned_query.append((key, value))

    query = urlencode(cleaned_query, doseq=True)
    return urlunparse((scheme, netloc, path, "", query, ""))


def job_identity_keys(record: Mapping[str, Any]) -> list[str]:
    keys: list[str] = []
    seen = set()

    for field_name in ("apply_link", "linkedin_link", "link", "source_url"):
        canonical_url = canonicalize_url(str(record.get(field_name) or ""))
        if canonical_url:
            identity = f"url:{canonical_url}"
            if identity not in seen:
                keys.append(identity)
                seen.add(identity)

    job_id = compact_whitespace(str(record.get("job_id") or ""))
    if job_id:
        identity = f"job_id:{job_id}"
        if identity not in seen:
            keys.append(identity)
            seen.add(identity)

    signature = title_company_signature(record.get("title"), record.get("company"))
    if signature:
        identity = f"title_company:{signature}"
        if identity not in seen:
            keys.append(identity)
            seen.add(identity)

    return keys


def dedupe_job_records(
    records: Iterable[Mapping[str, Any]],
    *,
    existing_keys: set[str] | None = None,
    logger: logging.Logger | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    active_logger = logger or LOGGER
    seen_keys = set(existing_keys or set())
    kept: list[dict[str, Any]] = []
    dropped: list[dict[str, Any]] = []

    for record in records:
        record_dict = dict(record)
        identity_keys = job_identity_keys(record_dict)
        duplicate_key = next((key for key in identity_keys if key in seen_keys), None)
        if duplicate_key:
            dropped_record = {
                **record_dict,
                "dedupe_reason": f"duplicate_identity:{duplicate_key}",
            }
            dropped.append(dropped_record)
            active_logger.info(
                "Skipping duplicate job '%s' at '%s' via %s",
                record_dict.get("title", ""),
                record_dict.get("company", ""),
                duplicate_key,
            )
            continue

        kept.append(record_dict)
        seen_keys.update(identity_keys)

    return kept, dropped


def load_existing_tracker_identity_keys(excel_path: str | Path) -> set[str]:
    path = Path(excel_path or "").expanduser()
    if not path.exists() or not path.is_file():
        return set()

    try:
        from openpyxl import load_workbook
    except Exception:
        LOGGER.warning("openpyxl is unavailable; tracker dedupe will skip workbook '%s'.", path)
        return set()

    identity_keys: set[str] = set()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            if not header_row:
                continue
            headers = [compact_whitespace(str(cell or "")) for cell in header_row]
            if not any(headers):
                continue

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                record = {headers[index]: row[index] for index in range(min(len(headers), len(row))) if headers[index]}
                identity_keys.update(job_identity_keys(record))
    finally:
        workbook.close()

    return identity_keys


__all__ = [
    "canonicalize_url",
    "compact_whitespace",
    "dedupe_job_records",
    "job_identity_keys",
    "load_existing_tracker_identity_keys",
    "normalize_title_company_part",
    "title_company_signature",
]
