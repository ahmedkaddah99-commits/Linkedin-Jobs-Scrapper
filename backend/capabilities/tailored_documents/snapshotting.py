import json
from pathlib import Path

from .common import compact_whitespace


def normalize_job_signature_part(value) -> str:
    return compact_whitespace(str(value or "")).lower()


def make_job_signature(title, company) -> str:
    title_part = normalize_job_signature_part(title)
    company_part = normalize_job_signature_part(company)
    if not title_part or not company_part:
        return ""
    return f"{title_part}||{company_part}"


def load_existing_job_signatures_from_excel(excel_path: str):
    path = Path(excel_path or "").expanduser()
    if not path.exists() or not path.is_file():
        return set()

    try:
        from openpyxl import load_workbook
    except Exception:
        print("[Stage1] warning: openpyxl not available; skipping Excel title+company prefilter.")
        return set()

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"[Stage1] warning: failed reading Excel '{path}': {exc}. Skipping Excel prefilter.")
        return set()

    existing_signatures = set()
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            header_row = next(rows, None)
            if not header_row:
                continue

            headers = [str(cell or "").strip().lower() for cell in header_row]
            if "title" not in headers or "company" not in headers:
                continue
            title_col_index = headers.index("title")
            company_col_index = headers.index("company")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if title_col_index >= len(row) or company_col_index >= len(row):
                    continue
                signature = make_job_signature(row[title_col_index], row[company_col_index])
                if signature:
                    existing_signatures.add(signature)
    finally:
        workbook.close()

    return existing_signatures


def load_jobs_snapshot(path_value: str):
    path = Path(path_value or "").expanduser()
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(path)
    with path.open("r", encoding="utf-8") as file:
        payload = json.load(file)
    if not isinstance(payload, list):
        raise ValueError(f"Snapshot must contain a list, got {type(payload).__name__}")
    return payload


def save_jobs_snapshot(path_value: str, jobs):
    path = Path(path_value or "").expanduser()
    with path.open("w", encoding="utf-8") as file:
        json.dump(jobs, file, indent=4, ensure_ascii=False)
