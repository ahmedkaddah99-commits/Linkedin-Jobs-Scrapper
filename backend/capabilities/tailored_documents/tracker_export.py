import json
from pathlib import Path
from typing import Dict, List


def to_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def derive_columns(records: List[Dict]) -> List[str]:
    preferred_columns = [
        "run_date",
        "run_timestamp",
        "job_id",
        "title",
        "company",
        "location_raw",
        "keyword",
        "posted_time_text",
        "posted_age_hours",
        "applicant_count",
        "priority_rank",
        "priority_rule",
        "easy_apply_status",
        "apply_link",
        "apply_link_source",
        "linkedin_link",
        "link",
        "enrich_status_code",
        "enrich_error",
        "full_description",
        "cv_professional_summary",
        "cv_professional_experience",
        "cv_strategic_initiatives",
        "cv_skills",
        "cv_education",
        "applied_cv",
        "tailored_cv",
        "cv_docx",
        "cv_pdf",
        "tailored_cv_docx",
        "pdf_generation_error",
        "doc_generation_error",
    ]

    seen = set()
    for record in records:
        seen.update(record.keys())

    columns = []
    for column in preferred_columns:
        if column in seen:
            columns.append(column)
            seen.discard(column)

    columns.extend(sorted(seen))
    return columns


def sheet_is_empty(worksheet) -> bool:
    return worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None


def add_sheet_with_unique_name(workbook, base_name: str):
    clean_base = (base_name or "jobs").strip() or "jobs"
    clean_base = clean_base[:28]
    candidate = clean_base
    counter = 2
    while candidate in workbook.sheetnames:
        suffix = f"_{counter}"
        candidate = f"{clean_base[:31-len(suffix)]}{suffix}"
        counter += 1
    return workbook.create_sheet(title=candidate)


def apply_hyperlink(cell, raw_value: str) -> None:
    value = (raw_value or "").strip()
    if not value:
        return

    if value.startswith("http://") or value.startswith("https://"):
        cell.hyperlink = value
        cell.style = "Hyperlink"
        return

    path = Path(value)
    if path.exists():
        cell.hyperlink = path.resolve().as_uri()
        cell.style = "Hyperlink"


def style_worksheet(worksheet, headers: List[str]) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    for header_cell in worksheet[1]:
        header_cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    width_defaults = {
        "run_date": 12,
        "run_timestamp": 22,
        "job_id": 14,
        "title": 42,
        "company": 32,
        "location_raw": 24,
        "keyword": 22,
        "posted_time_text": 20,
        "posted_age_hours": 14,
        "applicant_count": 14,
        "priority_rank": 12,
        "priority_rule": 34,
        "easy_apply_status": 16,
        "apply_link": 45,
        "apply_link_source": 18,
        "linkedin_link": 45,
        "link": 45,
        "enrich_status_code": 16,
        "enrich_error": 30,
        "full_description": 80,
        "cv_professional_summary": 80,
        "cv_professional_experience": 90,
        "cv_strategic_initiatives": 90,
        "cv_skills": 55,
        "cv_education": 80,
        "applied_cv": 45,
        "tailored_cv": 80,
        "cv_docx": 45,
        "cv_pdf": 45,
        "tailored_cv_docx": 45,
        "pdf_generation_error": 35,
        "doc_generation_error": 30,
    }

    for index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(index)
        worksheet.column_dimensions[column_letter].width = width_defaults.get(header, 24)

    wrap_columns = {
        "full_description",
        "cv_professional_summary",
        "cv_professional_experience",
        "cv_strategic_initiatives",
        "cv_skills",
        "cv_education",
        "applied_cv",
        "tailored_cv",
        "pdf_generation_error",
        "enrich_error",
        "doc_generation_error",
    }
    header_to_index = {header: idx for idx, header in enumerate(headers, start=1)}
    for column_name in wrap_columns:
        if column_name not in header_to_index:
            continue
        col_index = header_to_index[column_name]
        column_letter = get_column_letter(col_index)
        for cell in worksheet[column_letter]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_records_to_worksheet(worksheet, records: List[Dict], headers: List[str], append_only: bool) -> None:
    hyperlink_columns = {
        "apply_link",
        "linkedin_link",
        "link",
        "applied_cv",
        "cv_docx",
        "cv_pdf",
        "tailored_cv_docx",
    }
    header_to_index = {header: idx for idx, header in enumerate(headers, start=1)}

    if not append_only:
        worksheet.append(headers)

    for record in records:
        row_values = [to_cell_value(record.get(header)) for header in headers]
        worksheet.append(row_values)
        row_index = worksheet.max_row
        for column_name in hyperlink_columns:
            col_index = header_to_index.get(column_name)
            if not col_index:
                continue
            cell = worksheet.cell(row=row_index, column=col_index)
            apply_hyperlink(cell, cell.value or "")


def save_to_excel(records: List[Dict], output_path: Path, excel_mode: str, sheet_name: str, run_date: str) -> None:
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required for Excel export. Install with: pip install openpyxl") from exc

    if output_path.exists():
        workbook = load_workbook(output_path)
    else:
        workbook = Workbook()

    headers = derive_columns(records)

    if excel_mode == "append-rows":
        target_name = (sheet_name or "jobs").strip() or "jobs"
        if target_name in workbook.sheetnames:
            worksheet = workbook[target_name]
        else:
            worksheet = workbook.active if sheet_is_empty(workbook.active) else workbook.create_sheet(title=target_name)
            worksheet.title = target_name

        if sheet_is_empty(worksheet):
            write_records_to_worksheet(worksheet, records, headers, append_only=False)
        else:
            existing_headers = [cell.value for cell in worksheet[1] if cell.value]
            merged_headers = list(existing_headers)
            for header in headers:
                if header not in merged_headers:
                    merged_headers.append(header)
            if merged_headers != existing_headers:
                for index, header in enumerate(merged_headers, start=1):
                    worksheet.cell(row=1, column=index, value=header)
                headers = merged_headers
            else:
                headers = existing_headers
            write_records_to_worksheet(worksheet, records, headers, append_only=True)
    else:
        target_sheet_name = sheet_name or run_date
        if sheet_is_empty(workbook.active) and len(workbook.sheetnames) == 1:
            worksheet = workbook.active
            worksheet.title = (target_sheet_name[:31] or "jobs")
        else:
            worksheet = add_sheet_with_unique_name(workbook, target_sheet_name)
        write_records_to_worksheet(worksheet, records, headers, append_only=False)

    for ws in workbook.worksheets:
        current_headers = [cell.value for cell in ws[1] if cell.value]
        if current_headers:
            style_worksheet(ws, current_headers)

    workbook.save(output_path)
