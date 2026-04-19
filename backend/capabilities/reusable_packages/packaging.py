from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Dict, List, Mapping

from .support import cfg_str, compact_whitespace, load_json_file, load_reusable_packages_config, resolve_path, save_json_file


def sanitize_filename(value: str, max_length: int = 50) -> str:
    cleaned = "".join(char if char.isalnum() or char in ("-", "_", ".") else "_" for char in (value or ""))
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    cleaned = cleaned.strip("_")
    return (cleaned or "item")[:max_length]


def build_package_slug(job_id: str, company: str, title: str) -> str:
    base = sanitize_filename(f"{job_id}_{company}_{title}", max_length=44)
    if not base:
        base = sanitize_filename(job_id or "job", max_length=30)
    return base


def build_email_subject(job: Dict) -> str:
    title = compact_whitespace(str(job.get("title") or "Position"))
    return f"Bewerbung als {title} - Ahmed Kaddah"


def build_email_body(job: Dict, candidate_name: str, candidate_email: str, candidate_phone: str, category_name: str) -> str:
    company = compact_whitespace(str(job.get("company") or "Ihr Team"))
    title = compact_whitespace(str(job.get("title") or "die Position"))
    city = compact_whitespace(str(job.get("city") or job.get("location_raw") or ""))
    city_line = f" in {city}" if city else ""
    lines = [
        f"Sehr geehrtes Team von {company},",
        "",
        (
            f"hiermit bewerbe ich mich auf die Position {title}{city_line}. "
            f"Ich habe praktische Erfahrung in Logistik und operativen Teams und kann ab sofort in Vollzeit oder als Mini-Job starten."
        ),
        "",
        (
            f"Fuer diese Rolle bringe ich eine passende Hands-on-Erfahrung mit. "
            f"Ich habe mein Profil auf den Bereich '{category_name}' ausgerichtet und den entsprechenden Lebenslauf beigefuegt."
        ),
        "",
        "Ich freue mich ueber die Moeglichkeit eines persoenlichen Gespraechs.",
        "",
        "Mit freundlichen Gruessen",
        candidate_name,
    ]
    if candidate_email:
        lines.append(candidate_email)
    if candidate_phone:
        lines.append(candidate_phone)
    return "\n".join(lines).strip()


def load_role_cv_index(path: Path) -> Dict[str, Dict]:
    payload = load_json_file(path)
    rows = payload.get("role_cvs") if isinstance(payload, dict) else []
    mapping: Dict[str, Dict] = {}
    if isinstance(rows, list):
        for row in rows:
            if not isinstance(row, dict):
                continue
            category_id = str(row.get("category_id") or "").strip()
            if category_id:
                mapping[category_id] = row
    return mapping


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def write_excel(records: List[Dict], output_xlsx: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as exc:
        print(f"WARNING: openpyxl unavailable, skipping Excel export: {exc}")
        return

    headers = [
        "job_id",
        "portal",
        "title",
        "company",
        "location_raw",
        "role_category_id",
        "role_category_name",
        "classification_source",
        "link",
        "apply_link",
        "assigned_cv_txt",
        "assigned_cv_docx",
        "email_subject",
        "email_draft_path",
        "package_dir",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "job_packages"
    sheet.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for record in records:
        sheet.append([record.get(column, "") for column in headers])

    for column_cells in sheet.columns:
        max_len = 0
        for cell in column_cells:
            value = str(cell.value or "")
            if len(value) > max_len:
                max_len = len(value)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(70, max(12, max_len + 2))

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)


def build_stage5_args(config: dict | None = None, overrides: Mapping[str, Any] | None = None) -> SimpleNamespace:
    config = config or load_reusable_packages_config()
    payload = {
        "input": cfg_str(config, ("runtime", "stage5", "input_json"), "outputs/stage3_classified_jobs.json"),
        "role_cv_index_json": cfg_str(config, ("runtime", "stage5", "role_cv_index_json"), "outputs/stage4_role_cvs.json"),
        "output_json": cfg_str(config, ("runtime", "stage5", "output_json"), "outputs/stage5_application_packages.json"),
        "output_xlsx": cfg_str(config, ("runtime", "stage5", "output_xlsx"), "outputs/reusable_packages_with_docs.xlsx"),
        "docs_dir": cfg_str(config, ("runtime", "stage5", "docs_dir"), "outputs/generated_docs"),
        "run_date": "",
        "candidate_name": cfg_str(config, ("candidate", "name"), "Ahmed Kaddah"),
        "candidate_email": cfg_str(config, ("candidate", "email"), ""),
        "candidate_phone": cfg_str(config, ("candidate", "phone"), ""),
    }
    if overrides:
        payload.update({key: value for key, value in overrides.items() if value is not None})
    return SimpleNamespace(**payload)


def run_stage5_pipeline(
    args,
    *,
    config: dict | None = None,
    jobs: List[Dict] | None = None,
    role_index_payload: Dict[str, Any] | None = None,
) -> dict[str, Any]:
    _ = config
    input_path = resolve_path(args.input)
    role_index_path = resolve_path(args.role_cv_index_json)
    if jobs is None:
        if not input_path.exists():
            raise FileNotFoundError(f"input file not found: {input_path}")
        jobs = load_json_file(input_path)
        if not isinstance(jobs, list):
            raise ValueError("input JSON must be a list of jobs.")

    if role_index_payload is None:
        if not role_index_path.exists():
            raise FileNotFoundError(f"role CV index not found: {role_index_path}")
        role_index = load_role_cv_index(role_index_path)
    else:
        rows = role_index_payload.get("role_cvs") if isinstance(role_index_payload, dict) else []
        role_index = {}
        if isinstance(rows, list):
            for row in rows:
                if not isinstance(row, dict):
                    continue
                category_id = str(row.get("category_id") or "").strip()
                if category_id:
                    role_index[category_id] = row

    if not role_index:
        raise ValueError("role CV index is empty.")

    run_date = args.run_date.strip() or datetime.now().strftime("%Y-%m-%d")
    docs_root = ensure_dir(resolve_path(args.docs_dir) / run_date)
    records: List[Dict] = []

    for job in jobs:
        job_id = str(job.get("job_id") or "").strip() or "unknown"
        title = str(job.get("title") or "").strip() or "job"
        company = str(job.get("company") or "").strip() or "company"
        category_id = str(job.get("role_category_id") or "other_operational")
        category_name = str(job.get("role_category_name") or "Other Operational")
        role_cv = role_index.get(category_id) or role_index.get("other_operational") or {}

        safe_job_name = build_package_slug(job_id=job_id, company=company, title=title)
        package_dir = ensure_dir(docs_root / safe_job_name)

        role_cv_txt = Path(str(role_cv.get("cv_txt_path") or ""))
        role_cv_docx = Path(str(role_cv.get("cv_docx_path") or ""))
        assigned_cv_txt = ""
        assigned_cv_docx = ""
        if role_cv_txt.exists():
            assigned_cv_txt_path = package_dir / f"{safe_job_name}_CV.txt"
            shutil.copyfile(role_cv_txt, assigned_cv_txt_path)
            assigned_cv_txt = str(assigned_cv_txt_path)
        if role_cv_docx and str(role_cv_docx).strip() and role_cv_docx.exists():
            assigned_cv_docx_path = package_dir / f"{safe_job_name}_CV.docx"
            shutil.copyfile(role_cv_docx, assigned_cv_docx_path)
            assigned_cv_docx = str(assigned_cv_docx_path)

        email_subject = build_email_subject(job)
        email_body = build_email_body(job, args.candidate_name, args.candidate_email, args.candidate_phone, category_name)
        email_path = package_dir / f"{safe_job_name}_email.txt"
        email_path.write_text(f"Subject: {email_subject}\n\n{email_body}\n", encoding="utf-8")

        records.append(
            {
                **job,
                "run_date": run_date,
                "role_category_id": category_id,
                "role_category_name": category_name,
                "assigned_cv_txt": assigned_cv_txt,
                "assigned_cv_docx": assigned_cv_docx,
                "email_subject": email_subject,
                "email_draft_path": str(email_path),
                "package_dir": str(package_dir),
            }
        )

    output_json = resolve_path(args.output_json)
    output_xlsx = resolve_path(args.output_xlsx)
    save_json_file(output_json, records)
    write_excel(records, output_xlsx)
    print("Stage 5 complete.")
    print(f"Generated job packages: {len(records)}")
    print(f"Packages root: {docs_root}")
    print(f"Output JSON: {output_json}")
    print(f"Output XLSX: {output_xlsx}")
    return {
        "records": records,
        "output_json": str(output_json),
        "output_xlsx": str(output_xlsx),
        "docs_root": str(docs_root),
    }
