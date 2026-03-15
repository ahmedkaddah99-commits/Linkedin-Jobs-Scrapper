import argparse
import json
import os
import re
import time
from datetime import datetime, timedelta, timezone
from urllib.parse import parse_qs, quote, unquote, urlparse
from pathlib import Path

import requests as std_requests
from bs4 import BeautifulSoup
from scrapeops_python_requests.scrapeops_requests import ScrapeOpsRequests

from cv_profile import load_cv_text
from job_seeker_config import (
    cfg_bool,
    cfg_float,
    cfg_int,
    cfg_list,
    cfg_str,
    load_job_seeker_config,
    load_project_dotenv,
)


DEFAULT_KEYWORDS = [
    "analyst",
    "consultant",
    #"product manager",
    #"project manager",
    #"business process manager",
]

FORBIDDEN_WORDS = [
    "senior",
    "engineer",
    "sr",
    "sr.",
    "lead",
    "principal",
    "head",
    "director",
    "intern",
    "werkstudent",
]
DEFAULT_LOW_APPLICANT_THRESHOLD = int(os.getenv("STAGE1_LOW_APPLICANT_THRESHOLD", "80"))



def compact_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "")).strip()


def parse_posted_age_hours(posted_text: str):
    text = compact_whitespace((posted_text or "").lower())
    if not text:
        return None

    if "just now" in text or "today" in text:
        return 0.0
    if "yesterday" in text:
        return 24.0

    match = re.search(r"(\d+)\s*(minute|hour|day|week|month)s?", text)
    if not match:
        return None

    amount = int(match.group(1))
    unit = match.group(2)
    unit_hours = {
        "minute": 1 / 60,
        "hour": 1,
        "day": 24,
        "week": 24 * 7,
        "month": 24 * 30,
    }
    return round(amount * unit_hours.get(unit, 0), 2)


def extract_posted_time_text(soup: BeautifulSoup, page_text: str) -> str:
    selectors = [
        "span.posted-time-ago__text",
        "span.topcard__flavor--metadata",
        "span.topcard__flavor--bullet",
        "time",
    ]

    for selector in selectors:
        for node in soup.select(selector):
            text = compact_whitespace(node.get_text(separator=" ", strip=True))
            lowered = text.lower()
            if any(
                token in lowered
                for token in ["ago", "just now", "today", "yesterday", "reposted", "hour", "day", "week", "month"]
            ):
                return text

    fallback_match = re.search(
        r"(?i)(reposted\s+)?(\d+\s+(minute|hour|day|week|month)s?\s+ago|just now|today|yesterday)",
        page_text or "",
    )
    return compact_whitespace(fallback_match.group(0)) if fallback_match else ""


def extract_applicant_count(soup: BeautifulSoup, page_text: str):
    applicant_patterns = [
        re.compile(r"(?i)be among the first\s+(\d+)\s+applicants?"),
        re.compile(r"(?i)(?:over|more than)\s+(\d+)\+?\s+applicants?"),
        re.compile(r"(?i)(\d{1,3}(?:[,\.\s]\d{3})*|\d+)\+?\s+applicants?"),
    ]

    def parse_from_text(text: str):
        for pattern in applicant_patterns:
            match = pattern.search(text or "")
            if not match:
                continue
            digits = re.sub(r"[^\d]", "", match.group(1))
            if digits:
                return int(digits)
        return None

    selectors = [
        "figcaption.num-applicants__caption",
        "span.num-applicants__caption",
        "span.topcard__flavor--metadata",
    ]
    for selector in selectors:
        for node in soup.select(selector):
            count = parse_from_text(compact_whitespace(node.get_text(separator=" ", strip=True)))
            if count is not None:
                return count

    return parse_from_text(page_text or "")


def coerce_applicant_count(value):
    if value is None or value == "":
        return None
    try:
        return int(value)
    except Exception:
        digits = re.sub(r"[^\d]", "", str(value))
        return int(digits) if digits else None


def is_easy_apply(value) -> bool:
    return value is True


def priority_tier(job, low_applicant_threshold: int) -> int:
    easy = is_easy_apply(job.get("easy_apply_status"))
    applicants = coerce_applicant_count(job.get("applicant_count"))
    low_applicants = applicants is not None and applicants <= low_applicant_threshold

    if low_applicants and not easy:
        return 1
    if low_applicants and easy:
        return 2
    if not low_applicants and not easy:
        return 3
    return 4


def priority_sort_key(job, low_applicant_threshold: int):
    posted_age_hours = job.get("posted_age_hours")
    applicant_count = coerce_applicant_count(job.get("applicant_count"))

    posted_missing = posted_age_hours is None
    applicants_missing = applicant_count is None

    posted_value = float(posted_age_hours) if not posted_missing else float("inf")
    applicants_value = int(applicant_count) if not applicants_missing else 10**9

    return (
        priority_tier(job, low_applicant_threshold),
        1 if posted_missing else 0,
        posted_value,
        1 if applicants_missing else 0,
        applicants_value,
        str(job.get("job_id", "")),
    )


def normalize_job_signature_part(value) -> str:
    return compact_whitespace(str(value or "")).lower()


def make_job_signature(title, company) -> str:
    title_part = normalize_job_signature_part(title)
    company_part = normalize_job_signature_part(company)
    if not title_part or not company_part:
        return ""
    return f"{title_part}||{company_part}"


def load_existing_job_signatures_from_excel(excel_path: str):
    path = Path(excel_path or "").expanduser()
    if not path.exists() or not path.is_file():
        return set()

    try:
        from openpyxl import load_workbook
    except Exception:
        print("[Stage1] warning: openpyxl not available; skipping Excel title+company prefilter.")
        return set()

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"[Stage1] warning: failed reading Excel '{path}': {exc}. Skipping Excel prefilter.")
        return set()

    existing_signatures = set()
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            if not header_row:
                continue

            headers = [str(cell or "").strip().lower() for cell in header_row]
            if "title" not in headers or "company" not in headers:
                continue
            title_col_index = headers.index("title")
            company_col_index = headers.index("company")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if title_col_index >= len(row) or company_col_index >= len(row):
                    continue
                signature = make_job_signature(row[title_col_index], row[company_col_index])
                if signature:
                    existing_signatures.add(signature)
    finally:
        workbook.close()

    return existing_signatures


def load_jobs_snapshot(path_value: str):
    path = Path(path_value or "").expanduser()
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(path)
    with path.open("r", encoding="utf-8") as file:
        payload = json.load(file)
    if not isinstance(payload, list):
        raise ValueError(f"Snapshot must contain a list, got {type(payload).__name__}")
    return payload


def save_jobs_snapshot(path_value: str, jobs):
    path = Path(path_value or "").expanduser()
    with path.open("w", encoding="utf-8") as file:
        json.dump(jobs, file, indent=4, ensure_ascii=False)


def save_jobs_output(path_value: str, jobs) -> None:
    path = Path(path_value or "").expanduser()
    with path.open("w", encoding="utf-8") as file:
        json.dump(jobs, file, indent=4, ensure_ascii=False)


def extract_first_json_object(text: str) -> str:
    payload = text or ""
    start = payload.find("{")
    if start < 0:
        return ""

    depth = 0
    in_string = False
    escaped = False
    for index, char in enumerate(payload[start:], start=start):
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == '"':
                in_string = False
            continue

        if char == '"':
            in_string = True
            continue
        if char == "{":
            depth += 1
            continue
        if char == "}":
            depth -= 1
            if depth == 0:
                return payload[start : index + 1]

    return ""


def parse_ai_filter_payload(raw_content: str) -> dict:
    content = (raw_content or "").strip()
    cleaned = content.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
    candidates = [cleaned]

    extracted = extract_first_json_object(cleaned)
    if extracted and extracted != cleaned:
        candidates.append(extracted)

    for candidate in candidates:
        if not candidate:
            continue
        try:
            parsed = json.loads(candidate)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            continue

    approved_ids: list[str] = []
    approved_match = re.search(r'"approved_ids"\s*:\s*\[(.*?)\]', cleaned, re.DOTALL)
    if approved_match:
        approved_ids = re.findall(r'"(\d+)"|\b(\d+)\b', approved_match.group(1))
        approved_ids = [left or right for left, right in approved_ids if (left or right)]

    excluded_items = []
    excluded_match = re.search(r'"excluded"\s*:\s*\[(.*?)\]', cleaned, re.DOTALL)
    if excluded_match:
        for item_id, reason in re.findall(
            r'"id"\s*:\s*"?(.*?)"?\s*,\s*"reason"\s*:\s*"(.*?)"',
            excluded_match.group(1),
            re.DOTALL,
        ):
            excluded_items.append({"id": str(item_id).strip(), "reason": str(reason).strip()})

    if approved_ids or excluded_items:
        print("[Stage1] warning: recovered partial AI JSON using fallback parser.")
        return {"approved_ids": approved_ids, "excluded": excluded_items}

    preview = cleaned[:280].replace("\n", "\\n")
    raise ValueError(f"Unable to parse AI JSON payload. Preview={preview}")


def call_deepseek_title_filter(api_key: str, model: str, prompt: str):
    endpoint = "https://api.deepseek.com/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": "You are an expert career assistant."},
            {"role": "user", "content": prompt},
        ],
        "stream": False,
    }
    response = std_requests.post(endpoint, headers=headers, json=payload, timeout=120)
    response.raise_for_status()
    body = response.json()
    choices = body.get("choices") or []
    if not choices:
        raise ValueError("DeepSeek response missing choices")
    message = choices[0].get("message") or {}
    content = (message.get("content") or "").strip()
    return parse_ai_filter_payload(content)


def build_clients():
    load_project_dotenv()
    scrapeops_api_key = os.getenv("SCRAPEOPS_API_KEY")
    deepseek_api_key = os.getenv("DEEPSEEK_API_KEY")

    if not scrapeops_api_key or not deepseek_api_key:
        print(
            "ERROR: Missing API keys in environment/user_config/.env "
            "(SCRAPEOPS_API_KEY and DEEPSEEK_API_KEY are required)."
        )
        raise SystemExit(1)

    scrapeops_logger = ScrapeOpsRequests(
        scrapeops_api_key=scrapeops_api_key,
        spider_name="LinkedIn Smart Scraper",
        job_name="Germany-AI-Filtered",
    )
    so_requests = scrapeops_logger.RequestsWrapper()
    return scrapeops_api_key, deepseek_api_key, so_requests

def fetch_initial_list(
    keyword: str,
    so_requests,
    max_pages: int,
    geo_id: str,
    time_posted_seconds: int,
    experience_levels,
    forbidden_title_words,
    page_fetch_sleep_seconds: float,
):

    base_url = "https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search"
    found_jobs = []

    for page in range(max_pages):
        start_index = page * 25
        query_parts = [
            f"keywords={quote(keyword)}",
            f"geoId={quote(str(geo_id))}",
            f"start={start_index}",
        ]
        if int(time_posted_seconds) > 0:
            query_parts.append(f"f_TPR=r{int(time_posted_seconds)}")
        if experience_levels:
            exp_values = [str(int(level)) for level in experience_levels if str(level).strip()]
            if exp_values:
                query_parts.append(f"f_E={','.join(exp_values)}")
        query_params = "&".join(query_parts)
        target_url = f"{base_url}?{query_params}"

        print(f"[Stage1] {keyword} page {page + 1}: fetching")

        try:
            response = so_requests.get(target_url)
            if response.status_code != 200:
                break

            soup = BeautifulSoup(response.text, "html.parser")
            items = soup.find_all("li")
            if not items:
                break

            for item in items:
                card_text = item.get_text(separator=" ", strip=True).lower()
                if "reposted" in card_text:
                    continue

                base_card = item.find("div", {"class": "base-card"})
                if not base_card:
                    continue

                title_tag = item.find("h3", {"class": "base-search-card__title"})
                company_tag = item.find("h4", {"class": "base-search-card__subtitle"})
                location_tag = item.find("span", {"class": "job-search-card__location"})
                if not title_tag or not company_tag:
                    continue

                title = title_tag.get_text(strip=True)
                if any(word in title.lower() for word in forbidden_title_words):
                    continue

                company = company_tag.get_text(strip=True)
                location_raw = location_tag.get_text(strip=True) if location_tag else ""
                urn = base_card.get("data-entity-urn", "")
                if ":" not in urn:
                    continue

                job_id = urn.split(":")[-1]
                found_jobs.append(
                    {
                        "job_id": job_id,
                        "title": title,
                        "company": company,
                        "location_raw": location_raw,
                        "keyword": keyword,
                        "link": f"https://www.linkedin.com/jobs/view/{job_id}",
                        "linkedin_link": f"https://www.linkedin.com/jobs/view/{job_id}",
                    }
                )

            if page_fetch_sleep_seconds > 0:
                time.sleep(page_fetch_sleep_seconds)
        except Exception as exc:
            print(f"[Stage1] error for keyword={keyword} page={page + 1}: {exc}")
            break

    return found_jobs


def filter_with_ai(
    jobs_list,
    deepseek_api_key: str,
    cv_summary: str,
    model: str,
    excluded_output: str,
    extra_instructions: str = "",
    prompt_override: str = "",
):

    if not jobs_list:
        return [], []

    print("[Stage1] sending job titles to AI filter")

    job_inventory = "\n".join(
        [f"ID: {job['job_id']} | Title: {job['title']} | Company: {job['company']}" for job in jobs_list]
    )

    if prompt_override.strip():
        prompt = (
            prompt_override.strip()
            .replace("{{CV_SUMMARY}}", cv_summary)
            .replace("{{JOB_LIST}}", job_inventory)
        )
    else:
        prompt = f"""
You are an expert career assistant. I will give you my CV summary and a list of job titles.

MY CV SUMMARY:
{cv_summary}

JOB LIST:
{job_inventory}

YOUR TASK:
Evaluate every job in the list.

Rules:
- APPROVE a job ONLY IF:
  1) The title is written in English
  2) The title is relevant to my CV (Business Transformation, AI, Project/Product Management, Consulting, Data/Business Analysis)

OUTPUT FORMAT (IMPORTANT):
Return ONLY raw JSON (no markdown, no extra text), shaped EXACTLY like this:

{{
  "approved_ids": ["123", "456"],
  "excluded": [
    {{
      "id": "789",
      "reason": "German title"
    }},
    {{
      "id": "1011",
      "reason": "Not relevant"
    }},
    {{
      "id": "1213",
      "reason": "German title + Not relevant"
    }}
  ]
}}
""".strip()
    if extra_instructions:
        prompt = f"{prompt}\n\nAdditional user preferences:\n{extra_instructions.strip()}"

    parsed = None
    attempts = [
        prompt,
        (
            f"{prompt}\n\n"
            "Return strict JSON only. Ensure valid escaping for all strings. "
            "Do not include markdown fences or commentary."
        ),
    ]
    for attempt_index, attempt_prompt in enumerate(attempts, start=1):
        try:
            parsed = call_deepseek_title_filter(
                api_key=deepseek_api_key,
                model=model,
                prompt=attempt_prompt,
            )
            break
        except Exception as exc:
            if attempt_index == len(attempts):
                print(f"[Stage1] AI filtering failed: {exc}. Returning 0 approved jobs to protect credits.")
                return [], []
            print(f"[Stage1] AI JSON parse failed (attempt {attempt_index}/{len(attempts)}): {exc}. Retrying...")

    approved_ids = {str(item) for item in parsed.get("approved_ids", [])}
    excluded_map = {
        str(item.get("id")): item.get("reason", "Excluded")
        for item in parsed.get("excluded", [])
        if isinstance(item, dict) and item.get("id") is not None
    }

    approved_jobs = [job for job in jobs_list if str(job["job_id"]) in approved_ids]
    excluded_jobs = []
    for job in jobs_list:
        if str(job["job_id"]) not in approved_ids:
            excluded_jobs.append({**job, "reason": excluded_map.get(str(job["job_id"]), "Excluded by DeepSeek")})

    with open(excluded_output, "w", encoding="utf-8") as file:
        json.dump(excluded_jobs, file, indent=4, ensure_ascii=False)

    print(f"[Stage1] AI approved {len(approved_jobs)} / {len(jobs_list)}")
    print(f"[Stage1] wrote AI excluded jobs to {excluded_output}")
    return approved_jobs, excluded_jobs


def enrich_job(
    job_id: str,
    so_requests,
    scrapeops_api_key: str,
    debug_enrich_blocks: bool,
    use_proxy_fallback: bool,
):
    target_url = f"https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/{job_id}"

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }

    try:
        attempts = []
        resp = so_requests.get(target_url, headers=headers, timeout=45)
        attempts.append(("so_requests", resp.status_code))

        if resp.status_code != 200:
            resp = std_requests.get(target_url, headers=headers, timeout=45)
            attempts.append(("direct", resp.status_code))

        if resp.status_code != 200 and use_proxy_fallback:
            proxy_url = "https://proxy.scrapeops.io/v1/"
            params = {
                "api_key": scrapeops_api_key,
                "url": target_url,
                "residential": "true",
                "render_js": "true",
            }
            resp = std_requests.get(proxy_url, params=params, headers=headers, timeout=45)
            attempts.append(("scrapeops_proxy", resp.status_code))

        if resp.status_code != 200:
            if debug_enrich_blocks:
                print("\n===== ENRICH DEBUG (BLOCKED) =====")
                print("Status:", resp.status_code)
                print("Attempts:", attempts)
                print("Body preview:", resp.text[:300])
                print("==================================\n")

            return {
                "easy_apply_status": "unknown",
                "description": None,
                "apply_link": f"https://www.linkedin.com/jobs/view/{job_id}",
                "apply_link_source": "linkedin_fallback",
                "enrich_error": f"Failed (Status {resp.status_code}) | Attempts={attempts}",
                "status_code": resp.status_code,
            }

        html_lower = resp.text.lower()
        soup = BeautifulSoup(resp.text, "html.parser")
        page_text = compact_whitespace(soup.get_text(separator=" ", strip=True))

        primary_apply = soup.select_one(
            ".top-card-layout__cta--primary[data-tracking-control-name*='public_jobs_apply-link']"
        )
        apply_trackings = [
            (el.get("data-tracking-control-name") or "").lower()
            for el in soup.select("[data-tracking-control-name*='public_jobs_apply-link']")
        ]
        primary_text = primary_apply.get_text(separator=" ", strip=True).lower() if primary_apply else ""

        if "easy apply" in primary_text or "einfach bewerben" in primary_text:
            easy_apply_status = True
        elif any("easy_apply" in item for item in apply_trackings):
            easy_apply_status = True
        elif any("onsite" in item for item in apply_trackings):
            easy_apply_status = True
        elif any("offsite" in item for item in apply_trackings):
            easy_apply_status = False
        elif "easy apply" in html_lower or "einfach bewerben" in html_lower:
            easy_apply_status = True
        elif "apply on company website" in html_lower:
            easy_apply_status = False
        else:
            easy_apply_status = "unknown"

        desc_tag = (
            soup.find("div", {"class": "show-more-less-html__markup"})
            or soup.find("div", {"class": "description__text"})
            or soup.select_one("section.show-more-less-html")
        )
        description = desc_tag.get_text(separator="\n", strip=True) if desc_tag else None

        def normalize_href(raw_href: str):
            href = (raw_href or "").strip()
            if not href or href == "#" or href.lower().startswith("javascript:"):
                return None
            if href.startswith("//"):
                href = f"https:{href}"
            elif href.startswith("/"):
                href = f"https://www.linkedin.com{href}"
            return href

        def decode_session_redirect(url: str):
            try:
                parsed = urlparse(url)
                redirect_value = parse_qs(parsed.query).get("session_redirect", [None])[0]
                if redirect_value:
                    return unquote(redirect_value)
            except Exception:
                pass
            return url

        linkedin_default = f"https://www.linkedin.com/jobs/view/{job_id}"
        external_apply_link = None
        linkedin_apply_link = None

        apply_elements = soup.select("[data-tracking-control-name*='public_jobs_apply-link'], [data-apply-url], a[href]")
        for element in apply_elements:
            tracking = (element.get("data-tracking-control-name") or "").lower()
            href = normalize_href(
                element.get("href") or element.get("data-apply-url") or element.get("data-test-apply-url")
            )
            if not href:
                continue

            if "linkedin.com/login" in href.lower() or "/uas/login" in href.lower():
                href = decode_session_redirect(href)

            href_lower = href.lower()
            is_apply_candidate = (
                "apply" in tracking
                or "apply" in href_lower
                or "jobs/view" in href_lower
            )
            if not is_apply_candidate:
                continue

            if "linkedin.com" not in href_lower:
                external_apply_link = href
                break

            if "linkedin.com/jobs/view" in href_lower and not linkedin_apply_link:
                linkedin_apply_link = href

        if external_apply_link:
            apply_link = external_apply_link
            apply_link_source = "external"
        elif linkedin_apply_link:
            apply_link = linkedin_apply_link
            apply_link_source = "linkedin"
        else:
            apply_link = linkedin_default
            apply_link_source = "linkedin_fallback"

        posted_time_text = extract_posted_time_text(soup, page_text)
        posted_age_hours = parse_posted_age_hours(posted_time_text)
        applicant_count = extract_applicant_count(soup, page_text)
        posted_datetime_estimated_utc = None
        if posted_age_hours is not None:
            posted_datetime_estimated_utc = (
                datetime.now(timezone.utc) - timedelta(hours=float(posted_age_hours))
            ).isoformat(timespec="seconds")

        return {
            "easy_apply_status": easy_apply_status,
            "description": description,
            "apply_link": apply_link,
            "apply_link_source": apply_link_source,
            "posted_time_text": posted_time_text,
            "posted_age_hours": posted_age_hours,
            "posted_datetime_estimated_utc": posted_datetime_estimated_utc,
            "applicant_count": applicant_count,
            "enrich_error": None if description else "Description container not found",
            "status_code": resp.status_code,
        }
    except Exception as exc:
        return {
            "easy_apply_status": "unknown",
            "description": None,
            "apply_link": f"https://www.linkedin.com/jobs/view/{job_id}",
            "apply_link_source": "linkedin_fallback",
            "posted_time_text": "",
            "posted_age_hours": None,
            "posted_datetime_estimated_utc": None,
            "applicant_count": None,
            "enrich_error": f"Error: {exc}",
            "status_code": -1,
        }


def main() -> int:
    config = load_job_seeker_config()
    default_keywords = [str(item) for item in cfg_list(config, ("job_search", "keywords"), DEFAULT_KEYWORDS)]
    default_geo_id = cfg_str(
        config,
        ("job_search", "linkedin_geo_id"),
        os.getenv("LINKEDIN_GEO_ID", "101282230"),
    )
    default_time_posted_seconds = cfg_int(
        config,
        ("job_search", "time_posted_seconds"),
        int(os.getenv("STAGE1_TIME_POSTED_SECONDS", "86400")),
    )
    default_experience_levels = [
        int(level)
        for level in cfg_list(config, ("job_search", "experience_levels"), [2, 3])
        if str(level).strip()
    ]
    default_forbidden_title_words = [
        str(item)
        for item in cfg_list(config, ("job_search", "forbidden_title_keywords"), FORBIDDEN_WORDS)
        if str(item).strip()
    ]
    default_stage1_model = cfg_str(
        config,
        ("ai", "models", "stage1_title_filter_deepseek"),
        cfg_str(
            config,
            ("ai", "models", "stage1_title_filter"),
            os.getenv("DEEPSEEK_STAGE1_MODEL", "deepseek-chat"),
        ),
    )
    default_stage1_extra_prompt = cfg_str(config, ("ai", "prompts", "stage1_extra_instructions"), "")
    default_stage1_prompt_override = cfg_str(config, ("ai", "prompts", "stage1_prompt_override"), "")
    default_low_applicant_threshold = cfg_int(
        config,
        ("job_search", "priority", "low_applicant_threshold"),
        DEFAULT_LOW_APPLICANT_THRESHOLD,
    )
    default_stage1_max_pages = cfg_int(
        config,
        ("runtime", "stage1", "max_pages"),
        int(os.getenv("STAGE1_MAX_PAGES", "1")),
    )
    default_stage1_max_enrich_jobs = cfg_int(
        config,
        ("runtime", "stage1", "max_enrich_jobs"),
        int(os.getenv("STAGE1_MAX_ENRICH_JOBS", "30")),
    )
    default_stage1_output = cfg_str(
        config,
        ("runtime", "stage1", "output_json"),
        "highly_curated_jobs.json",
    )
    default_stage1_excluded_output = cfg_str(
        config,
        ("runtime", "stage1", "excluded_output_json"),
        "deepseek_excluded_jobs.json",
    )
    default_existing_jobs_excel = cfg_str(
        config,
        ("outputs", "stage4_xlsx"),
        "final_jobs_with_docs.xlsx",
    )
    default_stage1_scrape_snapshot_json = cfg_str(
        config,
        ("runtime", "stage1", "scrape_snapshot_json"),
        "stage1_scrape_snapshot.json",
    )
    default_stage1_reuse_scrape_snapshot = cfg_bool(
        config,
        ("runtime", "stage1", "reuse_scrape_snapshot"),
        os.getenv("STAGE1_REUSE_SCRAPE_SNAPSHOT", "false").lower() in ("1", "true", "yes"),
    )
    default_debug_enrich_blocks = cfg_bool(
        config,
        ("runtime", "stage1", "debug_enrich_blocks"),
        os.getenv("DEBUG_ENRICH_BLOCKS", "true").lower() in ("1", "true", "yes"),
    )
    default_page_fetch_sleep_seconds = cfg_float(
        config,
        ("runtime", "stage1", "page_fetch_sleep_seconds"),
        1.0,
    )
    default_use_proxy_fallback = cfg_bool(
        config,
        ("runtime", "stage1", "use_scrapeops_proxy_fallback"),
        os.getenv("USE_SCRAPEOPS_PROXY_STAGE4", "false").lower() in ("1", "true", "yes"),
    )

    parser = argparse.ArgumentParser(description="Stage 1: scrape, AI title-filter, and enrich jobs.")
    parser.add_argument("--max-pages", type=int, default=default_stage1_max_pages)
    parser.add_argument(
        "--max-enrich-jobs",
        type=int,
        default=default_stage1_max_enrich_jobs,
        help="Use -1 to enrich all approved jobs.",
    )
    parser.add_argument("--keywords", nargs="*", default=default_keywords)
    parser.add_argument("--geo-id", default=default_geo_id, help="LinkedIn geoId to search in.")
    parser.add_argument(
        "--time-posted-seconds",
        type=int,
        default=default_time_posted_seconds,
        help="LinkedIn time-posted filter in seconds (0 disables filter).",
    )
    parser.add_argument(
        "--experience-levels",
        nargs="*",
        type=int,
        default=default_experience_levels,
        help="LinkedIn experience level codes, e.g. 2 3.",
    )
    parser.add_argument(
        "--forbidden-title-keywords",
        nargs="*",
        default=default_forbidden_title_words,
        help="Title keywords to exclude before AI filtering.",
    )
    parser.add_argument("--output", default=default_stage1_output)
    parser.add_argument("--excluded-output", default=default_stage1_excluded_output)
    parser.add_argument(
        "--existing-jobs-excel",
        default=default_existing_jobs_excel,
        help="Excel file path used to skip jobs whose title+company already exist across any sheet before AI filtering.",
    )
    parser.add_argument(
        "--model",
        default=default_stage1_model,
        help="DeepSeek model for Stage 1 title filtering (e.g., deepseek-chat).",
    )
    parser.add_argument(
        "--scrape-snapshot-json",
        default=default_stage1_scrape_snapshot_json,
        help="Path to Stage 1 scrape snapshot (saved before AI filtering).",
    )
    parser.add_argument(
        "--reuse-scrape-snapshot",
        action=argparse.BooleanOptionalAction,
        default=default_stage1_reuse_scrape_snapshot,
        help="Skip scraping and reuse jobs from --scrape-snapshot-json.",
    )
    parser.add_argument(
        "--stage1-extra-prompt",
        default=default_stage1_extra_prompt,
        help="Extra instructions appended to Stage 1 AI title-filter prompt.",
    )
    parser.add_argument(
        "--stage1-prompt-override",
        default=default_stage1_prompt_override,
        help="Optional full prompt override. Supports {{CV_SUMMARY}} and {{JOB_LIST}} placeholders.",
    )
    parser.add_argument(
        "--low-applicant-threshold",
        type=int,
        default=default_low_applicant_threshold,
        help="Applicant count threshold separating low vs high applicant groups for prioritization.",
    )
    parser.add_argument(
        "--debug-enrich-blocks",
        action=argparse.BooleanOptionalAction,
        default=default_debug_enrich_blocks,
    )
    parser.add_argument(
        "--page-fetch-sleep-seconds",
        type=float,
        default=default_page_fetch_sleep_seconds,
        help="Sleep between paginated LinkedIn requests (seconds).",
    )
    parser.add_argument(
        "--use-proxy-fallback",
        action=argparse.BooleanOptionalAction,
        default=default_use_proxy_fallback,
        help="Use ScrapeOps proxy fallback when direct enrichment fails.",
    )
    args = parser.parse_args()

    scrapeops_api_key, deepseek_api_key, so_requests = build_clients()
    cv_summary = load_cv_text()

    print("[Stage1] starting pipeline")
    forbidden_title_words = [str(item).lower() for item in (args.forbidden_title_keywords or []) if str(item).strip()]
    if not forbidden_title_words:
        forbidden_title_words = [str(item).lower() for item in FORBIDDEN_WORDS]

    if args.reuse_scrape_snapshot:
        try:
            unique_jobs_list = load_jobs_snapshot(args.scrape_snapshot_json)
            print(
                f"[Stage1] reusing scrape snapshot: {args.scrape_snapshot_json} "
                f"({len(unique_jobs_list)} jobs)"
            )
        except Exception as exc:
            print(f"[Stage1] failed to load scrape snapshot '{args.scrape_snapshot_json}': {exc}")
            return 1
    else:
        all_raw_jobs = []
        for keyword in args.keywords:
            all_raw_jobs.extend(
                fetch_initial_list(
                    keyword=keyword,
                    so_requests=so_requests,
                    max_pages=args.max_pages,
                    geo_id=args.geo_id,
                    time_posted_seconds=max(0, int(args.time_posted_seconds)),
                    experience_levels=args.experience_levels or [],
                    forbidden_title_words=forbidden_title_words,
                    page_fetch_sleep_seconds=max(0.0, float(args.page_fetch_sleep_seconds)),
                )
            )

        unique_jobs_list = list({job["job_id"]: job for job in all_raw_jobs}.values())
        print(f"[Stage1] unique jobs found: {len(unique_jobs_list)}")
        try:
            save_jobs_snapshot(args.scrape_snapshot_json, unique_jobs_list)
            print(f"[Stage1] saved scrape snapshot: {args.scrape_snapshot_json}")
        except Exception as exc:
            print(f"[Stage1] warning: failed to save scrape snapshot '{args.scrape_snapshot_json}': {exc}")

    if not unique_jobs_list:
        print("[Stage1] no jobs available before AI filter, exiting.")
        save_jobs_output(args.output, [])
        print(f"[Stage1] wrote empty output to {args.output}")
        return 0

    existing_job_signatures = load_existing_job_signatures_from_excel(args.existing_jobs_excel)
    if existing_job_signatures:
        before_excel_prefilter = len(unique_jobs_list)
        unique_jobs_list = [
            job
            for job in unique_jobs_list
            if make_job_signature(job.get("title", ""), job.get("company", "")) not in existing_job_signatures
        ]
        skipped_count = before_excel_prefilter - len(unique_jobs_list)
        print(
            f"[Stage1] Excel prefilter: skipped {skipped_count} jobs with duplicate title+company found in "
            f"{args.existing_jobs_excel}"
        )
    else:
        print(f"[Stage1] Excel prefilter: no existing title+company pairs found in {args.existing_jobs_excel}")

    if not unique_jobs_list:
        print("[Stage1] no new jobs left after Excel prefilter, exiting.")
        save_jobs_output(args.output, [])
        print(f"[Stage1] wrote empty output to {args.output}")
        return 0

    ai_approved_jobs, _ = filter_with_ai(
        jobs_list=unique_jobs_list,
        deepseek_api_key=deepseek_api_key,
        cv_summary=cv_summary,
        model=args.model,
        excluded_output=args.excluded_output,
        extra_instructions=args.stage1_extra_prompt,
        prompt_override=args.stage1_prompt_override,
    )
    if not ai_approved_jobs:
        print("[Stage1] no jobs passed AI title filter, exiting.")
        save_jobs_output(args.output, [])
        print(f"[Stage1] wrote empty output to {args.output}")
        return 0

    if args.max_enrich_jobs >= 0:
        ai_approved_jobs = ai_approved_jobs[: args.max_enrich_jobs]
        print(f"[Stage1] enrichment capped to: {len(ai_approved_jobs)}")

    final_output = []
    total = len(ai_approved_jobs)
    print(f"[Stage1] enriching {total} jobs for easy-apply + full description")
    for index, job in enumerate(ai_approved_jobs, start=1):
        print(f"[Stage1] [{index}/{total}] {job['title']}")
        enrich = enrich_job(
            job_id=job["job_id"],
            so_requests=so_requests,
            scrapeops_api_key=scrapeops_api_key,
            debug_enrich_blocks=args.debug_enrich_blocks,
            use_proxy_fallback=args.use_proxy_fallback,
        )
        job["easy_apply_status"] = enrich["easy_apply_status"]
        job["full_description"] = enrich["description"]
        job["apply_link"] = enrich["apply_link"]
        job["apply_link_source"] = enrich["apply_link_source"]
        job["posted_time_text"] = enrich["posted_time_text"]
        job["posted_age_hours"] = enrich["posted_age_hours"]
        job["posted_datetime_estimated_utc"] = enrich["posted_datetime_estimated_utc"]
        job["applicant_count"] = enrich["applicant_count"]
        job["enrich_error"] = enrich["enrich_error"]
        job["enrich_status_code"] = enrich["status_code"]
        final_output.append(job)

        jd_ok = "OK" if job["full_description"] else "NO"
        print(
            f"[Stage1] result: EasyApply={job['easy_apply_status']} "
            f"| JD={jd_ok} | HTTP={job['enrich_status_code']}"
        )

    final_output.sort(key=lambda item: priority_sort_key(item, max(0, args.low_applicant_threshold)))
    for rank, job in enumerate(final_output, start=1):
        tier = priority_tier(job, max(0, args.low_applicant_threshold))
        tier_label = {
            1: "tier1_newest_non_easy_low_applicants",
            2: "tier2_newest_easy_low_applicants",
            3: "tier3_newest_non_easy_high_or_unknown_applicants",
            4: "tier4_newest_easy_high_or_unknown_applicants",
        }[tier]
        job["priority_rank"] = rank
        job["priority_tier"] = tier
        job["priority_bucket"] = tier_label
        job["priority_rule"] = (
            "1)newest+non_easy+low_applicants,"
            "2)newest+easy+low_applicants,"
            "3)newest+non_easy+high_or_unknown_applicants,"
            "4)newest+easy+high_or_unknown_applicants"
        )

    save_jobs_output(args.output, final_output)

    print(f"[Stage1] done. saved {len(final_output)} jobs to {args.output}")
    print(f"[Stage1] excluded jobs file: {args.excluded_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
